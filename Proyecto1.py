import string
import openpyxl
import os

l = ['o', 'k', ';', 'l', '.']
s = ['w', 'a', 's', 'd', 'x']

d = ['s', 'e', 'd', 's', 'f', 'c']
f = ['f', 'd', 'r', 'v', 'g']

p = ['p', 'o']
i2 = ['i', 'u', 'o', 'k']
n = ['n', 'm']
g = ['g', 'f', 'g', 't']


loc = ("./commands.xlsx")

if(os.path.exists(loc)):
    wb = openpyxl.load_workbook(loc)
else:
    wb = openpyxl.Workbook()
    wb.save(loc)

commands = ['ls', 'df', "ping"]

sheet = wb.active

for i in range(sheet.max_row):
    tem = sheet.cell(i+1, 1).value
    if tem:
        commands.append(tem)

max = sheet.max_row

while True:

    userInput = input(">")
    if(userInput == "exit"):
        wb.save(loc)
        print("Exec Finished")
        break

    if(len(userInput) > 1):
        if (userInput[:2] in commands or userInput[:4] in commands):
            if((userInput[0] in l and userInput[1] in s) or (userInput[0] in s and userInput[1] in l)):
                os.system("Dir")
            elif((userInput[0] in d and userInput[1] in f) or (userInput[0] in f and userInput[1] in d)):
                print("Disk Free: 234.43 GB")
            elif((userInput[0] in p and userInput[1] in i2 and userInput[2] in n and userInput[3] in g)):
                print("Ping to 192.168.1.2")
                

        else:

            if(userInput[0] in l and userInput[1] in s):
                if(userInput[1] in s):
                    wdym = input("Did you mean ls? Y/N")
                    if(wdym == 'Y' or wdym == 'y'):
                        commands.append(userInput[:2])
                        max = max + 1
                        sheet.cell(max, 1).value = userInput[:2]
                        os.system("Dir")
                else:
                    print("Command Unknown")

            elif(userInput[0] in s and userInput[1] in l):
                if(userInput[1] in l):
                    wdym = input("Did you mean ls? Y/N")
                    if(wdym == 'Y' or wdym == 'y'):
                        commands.append(userInput[:2])
                        max = max + 1
                        sheet.cell(max, 1).value = userInput[:2]
                        os.system("Dir")
                else:
                    print("Command Unknown")

            elif(userInput[0] in d):
                if (userInput[1] in f):
                    res = input("Did you mean df ? (Y/N)")
                    if(res == 'Y' or res == 'y'):
                        commands.append(userInput[:2])
                        max = max + 1
                        sheet.cell(max, 1).value = userInput[:2]
                        print("Disk Free: 234.43 GB")
                else:
                    print("Command Unknown")

            elif(userInput[0] in f):
                if (userInput[1] in d):
                    res = input("Did you mean df ? (Y/N)")
                    if(res == 'Y' or res == 'y'):
                        commands.append(userInput[:2])
                        max = max + 1
                        sheet.cell(max, 1).value = userInput[:2]
                        print("Disk Free: 234.43 GB")
                else:
                    print("Command Unknown")

            elif(userInput[0] in p):
                    if(userInput[1] in i2):
                        if(userInput[2] in n):
                            if(userInput[3] in g):
                                res = input("Did you mean ping ? (Y/N)")
                                if(res == 'Y' or res == 'y'):
                                    commands.append(userInput[:4])
                                    max = max + 1
                                    sheet.cell(max, 1).value = userInput[:4]
                                    print("Ping 192.168.1.2")
                            else:
                                print("Command Unknown")
